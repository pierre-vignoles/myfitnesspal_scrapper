import myfitnesspal
import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter
from my_config import *
import os
import pandas as pd


def function_scraping(date_scrapping: datetime) -> List[List[str]]:
    food_list = []
    food_list_final = []
    for meal in day.meals:
        name_meal = meal.name
        for al in meal:
            food_list.append(date_scrapping.strftime("%d/%m/%Y"))
            food_list.append(name_meal)
            food_list.append(al.name)
            food_list.append(al['calories'])
            food_list.append(al['protein'])
            food_list.append(al['fat'])
            food_list.append(al['carbohydrates'])
            food_list.append(al['sodium'])
            food_list.append(al['sugar'])

            food_list_final.append(food_list)
            food_list = []

    return food_list_final


def function_write_excel(food_list_final: List[List[str]], username_friend: str):
    # Create excel with column's name if not exist
    if os.path.isfile(path_excel / (name_file_excel + "_" + username_friend + ".xlsx")):
        excel_doc = openpyxl.load_workbook(path_excel / (name_file_excel + "_" + username_friend + ".xlsx"))
    else:
        excel_doc = openpyxl.Workbook()
        sheet = excel_doc.active
        sheet.title = name_sheet
        sheet['A1'] = name_column[0]
        sheet['B1'] = name_column[1]
        sheet['C1'] = name_column[2]
        sheet['D1'] = name_column[3]
        sheet['E1'] = name_column[4]
        sheet['F1'] = name_column[5]
        sheet['G1'] = name_column[6]
        sheet['H1'] = name_column[7]
        sheet['I1'] = name_column[8]

    sheet = excel_doc[name_sheet]
    max_row = sheet.max_row

    # Delete rows with the same date
    delete_rows: bool = False
    for rowNum in range(max_row, 1, -1):
        if sheet.cell(rowNum, 1).value == date_script.strftime("%d/%m/%Y"):
            first_row_idx = rowNum
            delete_rows: bool = True
            sheet.delete_rows(rowNum, 1)

    # Search the index line to write the new ones
    if delete_rows == True and food_list_final:
        sheet.insert_rows(first_row_idx, len(food_list_final))
    elif delete_rows == False and food_list_final:
        idx_row_to_write = sheet.max_row + 1
        for rowNum in range(2, max_row):
            if datetime.strptime(sheet.cell(rowNum, 1).value, "%d/%m/%Y").date() > date_script:
                idx_row_to_write = rowNum
                break
        sheet.insert_rows(idx_row_to_write, len(food_list_final))
    else:
        print("Nothing to write")

    # Write the new rows
    for idx_al, food in enumerate(food_list_final):
        for idx_data, data in enumerate(food):
            if delete_rows:
                sheet[str(get_column_letter(idx_data + 1)) + str(first_row_idx + idx_al)] = data
            else:
                sheet[str(get_column_letter(idx_data+1)) + str(idx_row_to_write + idx_al)] = data

    print("Writting the excel file : " + str(name_file_excel + "_" + username_friend))
    if safe_mode == True:
        excel_doc.save(path_excel / (name_file_excel + "_" + username_friend + "_" + date_script.strftime("%Y_%m_%d") + ".xlsx"))
    else:
        excel_doc.save(path_excel / (name_file_excel + "_" + username_friend + ".xlsx"))


if __name__ == '__main__':
    if manual_date_mode == True:
        if len(manual_date) == 1:
            date_script_array = manual_date
            print("Manual date mode activated. Date : " + str(date_script_array[0].strftime("%d/%m/%Y")))
        else:
            date_script_array = pd.date_range(manual_date[0], manual_date[1], freq='d')
            print("Manual date mode activated. Date : " + str(manual_date[0].strftime("%d/%m/%Y")) + " - "
                  + str(manual_date[1].strftime("%d/%m/%Y")))
    else:
        date_script_array = [datetime.now().date()]
        print("Manual date mode disabled. Date : " + str(date_script_array[0].strftime("%d/%m/%Y")))

    print("Connexion with myfitnesspal..")
    client = myfitnesspal.Client(username, password=password)

    for date_script in date_script_array:
        for username_friend in username_friend_list:
            print("\nScrape user : " + str(username_friend))
            if len(date_script_array) > 1:
                print("Date : " + str(date_script.strftime("%d/%m/%Y")))
            day = client.get_date(date_script.year, date_script.month, date_script.day, username=username_friend)
            food_list_final = function_scraping(date_script)

            function_write_excel(food_list_final, username_friend)

    print("\nEND")




